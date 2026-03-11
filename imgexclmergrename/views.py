
# Create your views here.
import os, json, zipfile
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from openpyxl import load_workbook
from difflib import SequenceMatcher

def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def index(request):
    if request.method == "POST" and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        action = request.POST.get("action")
        
        if action == "upload_excel":
            excel_file = request.FILES.get("excel_file")
            os.makedirs(os.path.join(settings.MEDIA_ROOT, "excel"), exist_ok=True)
            excel_path = os.path.join(settings.MEDIA_ROOT, "excel", excel_file.name)
            with open(excel_path,'wb+') as f:
                for chunk in excel_file.chunks():
                    f.write(chunk)
            wb = load_workbook(excel_path)
            ws = wb.active
            excel_names = [str(cell.value).strip() for cell in ws['A'] if cell.value]
            return JsonResponse({"excel_filename":excel_file.name,"total_items":len(excel_names),"excel_names":excel_names})
        
        elif action == "upload_images":
            os.makedirs(os.path.join(settings.MEDIA_ROOT, "images"), exist_ok=True)
            images = request.FILES.getlist("images")
            image_urls=[]
            for img in images:
                path=os.path.join(settings.MEDIA_ROOT,"images",img.name)
                with open(path,"wb+") as f:
                    for chunk in img.chunks():
                        f.write(chunk)
                image_urls.append(settings.MEDIA_URL+"images/"+img.name)
            return JsonResponse({"images":image_urls,"total_images":len(image_urls)})

        elif action=="comparison":
            images=json.loads(request.POST.get("images"))
            excel_names=json.loads(request.POST.get("excel_names"))
            matched=[]
            for i,img in enumerate(images):
                img_name=img.split("/")[-1]
                if i < len(excel_names):
                    score=similarity(img_name,excel_names[i])
                    if score>=0.5:
                        matched.append({"original":img_name,"renamed":excel_names[i],"original_url":img})
            return JsonResponse({"matched":matched})

        elif action=="rename_images":
            matched=json.loads(request.POST.get("matched"))
            renamed_dir=os.path.join(settings.MEDIA_ROOT,"renamed_images")
            os.makedirs(renamed_dir,exist_ok=True)
            renamed_images=[]
            for item in matched:
                original=item["original"]
                new_name=item["renamed"]
                old_path=os.path.join(settings.MEDIA_ROOT,"images",original)
                name,ext=os.path.splitext(original)
                new_filename=f"{new_name}{ext}"
                new_path=os.path.join(renamed_dir,new_filename)
                if os.path.exists(old_path):
                    os.rename(old_path,new_path)
                    renamed_images.append({"url":settings.MEDIA_URL+"renamed_images/"+new_filename,"name":new_filename})
            return JsonResponse({"success":True,"renamed_images":renamed_images})

        elif action=="download_folder":
            folder=os.path.join(settings.MEDIA_ROOT,"renamed_images")
            zip_path=os.path.join(settings.MEDIA_ROOT,"renamed_images.zip")
            with zipfile.ZipFile(zip_path,"w") as zipf:
                for root,dirs,files in os.walk(folder):
                    for file in files:
                        zipf.write(os.path.join(root,file),arcname=file)
            with open(zip_path,"rb") as f:
                response=HttpResponse(f.read(),content_type="application/zip")
                response['Content-Disposition']='attachment; filename=renamed_images.zip'
                return response

    return render(request,"index.html")